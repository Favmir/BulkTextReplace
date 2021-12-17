import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl.utils.cell
import re
import glob
import Tkinter as tk
import os

CONTENT_LOC = os.getcwd()
WORKBOOK_FILENAME = 'BulkReplacer_List.xlsx'
WORKBOOK_LOC = os.getcwd()
WORKBOOK_PATH =  os.path.join(WORKBOOK_LOC, WORKBOOK_FILENAME)

def GenerateFile():
    print('def GenerateFile() running')
    # Create "ReplaceList.xlsx" example file
    wb = Workbook()
    ws = wb.active
    ws.title = "Replace Text List"
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws['A1'] = 'Hi! This text will be replaced by'
    ws['B1'] = 'this text!'
    ws['C1'] = '<- This is the first example.'

    ws['A2'] = 'You can also replace using'
    ws['B2'] = 'multiple clauses!'
    ws['C2'] = '<- And this is the second example!'
    
    wb.save(WORKBOOK_PATH)

def OpenFile():
    if(os.path.exists(WORKBOOK_PATH) == False):
        GenerateFile()
    os.startfile(WORKBOOK_PATH)

def ReplaceText(ws):
    last_filled_row = len(list(ws.rows))
    for i in range(1, last_filled_row+1):
        print(i, ws['A'+str(i)].value, ws['B'+str(i)].value, ws['C'+str(i)].value)

    for filename in glob.glob('*.txt'):
        with open(os.path.join(CONTENT_LOC, filename), 'r+') as f:
            content = f.read()
            for numrow, row in enumerate(ws.iter_rows(), start = 1):
                content = re.sub(
                    row[0].internal_value,
                    row[1].internal_value,
                    content,
                    flags = re.M
                    )
            
            f.seek(0)
            f.write(content)
            f.truncate()

            f.close()


# Load workbook (needed for assgining commands to gui buttons)
if(os.path.exists(WORKBOOK_PATH) == False):
    GenerateFile
wb = load_workbook(filename = WORKBOOK_PATH)
ws = wb.active          # Crashes if you try non-exsisting sheet name workbook['Sheet1']

# openpyxl.utils.cell.get_column_letter
# xy = coordinate_from_string('A4') # returns ('A',4)
# col = column_index_from_string(xy[0]) # returns 1
# row = xy[1]

# Gui
window = tk.Tk()
frame1 = tk.Frame(master=window, width=300, height=150)
frame1.pack()
frame2 = tk.Frame(master=window, width=300, height=150)
frame2.pack()

lbl_desc1 = tk.Label(master=frame1, text='RegEx Bulk Replacer')
lbl_desc1.pack()

btn_openSheet = tk.Button(
    master = frame1,
    text = 'Open ' + WORKBOOK_FILENAME,
    width = 25,
    height = 5,
    command = OpenFile,
)
btn_openSheet.pack()

btn_preview = tk.Button(
    master = frame2,
    text = 'Preview changes',
    width = 25,
    height = 5
)
btn_preview.pack()

btn_run = tk.Button(
    master = frame2,
    text = 'Replace in every *.txt files',
    width = 25,
    height = 5,
    command = lambda: ReplaceText(ws)
)
btn_run.pack()

window.mainloop()








